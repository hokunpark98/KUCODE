from django.shortcuts import render
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.core.exceptions import ObjectDoesNotExist
from django.contrib.postgres.aggregates import ArrayAgg
from django.db.models import F, Q, Sum, Count, Avg, FloatField, IntegerField, Prefetch
from django.db.models.functions import Cast
from django.conf import settings

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from datetime import datetime
from account.models import User,Student,Administration
from course.models import Course, Course_registration, Course_project
from repo.models import Repo_commit, Repo_pr ,Repo_issue, Repository,Repo_contributor
from itertools import groupby
from operator import itemgetter
from collections import defaultdict
import statistics
import numpy as np
from openpyxl.styles import Alignment
from io import BytesIO
import openpyxl
from django.http import HttpResponse
from openpyxl.styles import Font
import io

import json
from django.http import JsonResponse

import requests
from account.api.views import get_kuopenapi_access_token,query_student_openapi,student_read_course_info

class HealthCheckAPIView(APIView):
    def get(self, request):
        return Response({"status": "OK"}, status=status.HTTP_200_OK)

# ========================================
# Frontend Function
# ========================================
def course_create_db (request):
    try:
        course_id=request.GET.get('course_id')
        year = request.GET.get('year')
        semester = request.GET.get('semester')
        name = request.GET.get('name')
        prof = request.GET.get('prof')
        ta = request.GET.get('ta')
        student_count = request.GET.get('student_count')

        if not course_id or not year or not semester:
            raise ValueError("course_id and year and semester cannot be empty")
        
        if Course.objects.filter(course_id=course_id, year=year,semester=semester).exists():
            raise ValueError("course already exists.")
        
        course = Course.objects.create(
            course_id=course_id,
            year=year,
            semester=semester,
            name=name,
            prof=prof,
            ta=ta,
            student_count=student_count
        )

        return JsonResponse({"status": "OK", "message": "course record created successfully"})
    
    except Exception as e:
        return JsonResponse({"status": "Error", "message": str(e)}, status=500)     

def course_read_db(request):
    try:
        # 1) 필요한 모든 데이터를 한 번에 가져오기 (select_related, prefetch_related)
        courses = Course.objects.select_related().prefetch_related(
            Prefetch(
                'course_project_set',
                queryset=Course_project.objects.select_related('repo').filter(
                    repo__forked=False
                )
            ),
            Prefetch(
                'course_registration_set',
                queryset=Course_registration.objects.select_related('student')
            )
        ).all()
        
        data = []
        
        for course in courses:
            # 2) 이미 prefetch된 데이터 사용 (추가 쿼리 없음)
            course_projects = course.course_project_set.all()
            
            # 3) DB aggregation 활용
            stats = course_projects.aggregate(
                total_commits=Sum('repo__commit_count'),
                total_issues=Sum(
                    F('repo__open_issue_count') + F('repo__closed_issue_count')
                ),
                total_prs=Sum(
                    F('repo__open_pr_count') + F('repo__closed_pr_count')
                ),
                total_stars=Sum('repo__star_count'),
                repo_count=Count('id')
            )
            # 학생 수 계산 방법 변경 (기존: course_course_registration의 course 별 수를 카운트하는 방식 -> 현재: course_course에서 학생 수를 직접 읽는 방식)
            # student_count = course.course_registration_set.count()
            student_count = course.student_count
            
            # contributor 계산 (이 부분만 Python 처리)
            contributor_count = 0
            for cp in course_projects:
                if cp.repo.contributors:
                    contributor_count += cp.repo.contributors.count(',') + 1
            
            data.append({
                "course_id": course.course_id,
                "year": course.year,
                "semester": course.semester,
                "name": course.name,
                "prof": course.prof,
                "ta": course.ta,
                "student_count": student_count,
                "total_commits": stats['total_commits'] or 0,
                "total_issues": stats['total_issues'] or 0,
                "total_prs": stats['total_prs'] or 0,
                "total_stars": stats['total_stars'] or 0,
                "avg_commits": round((stats['total_commits'] or 0) / student_count, 2) if student_count > 0 else 0,
                "repository_count": stats['repo_count'] or 0,
                "contributor_count": contributor_count
            })
        
        return JsonResponse(data, safe=False)
    
    except Exception as e:
        return JsonResponse({"status": "Error", "message": str(e)}, status=500)

def course_update_db(request):
    try:
        return JsonResponse({"status": "OK", "message": "course record update successfully"})
    except Exception as e:
        return JsonResponse({"status": "Error", "message": str(e)}, status=500)       

def course_delete_db(request):
    try:
        course_id=request.GET.get('course_id')
        year = request.GET.get('year')
        semester = request.GET.get('semester')           
        course = Course.objects.get(course_id=course_id, year=year, semester=semester)
        course.delete()  # 사용자 객체를 삭제합니다.
        return JsonResponse({"status": "OK", "message": "course record deleted successfully"})
    
    except ObjectDoesNotExist:
        return JsonResponse({"status": "Error", "message": f"Course '{course_id}' does not exist"}, status=404)    
    except Exception as e:
        return JsonResponse({"status": "Error", "message": str(e)}, status=500)



def course_registration_create_db(request):
    
    try:
        course_id = request.GET.get('course_id')
        year = request.GET.get('year')
        semester = request.GET.get('semester')  
        student_ids = Student.objects.values_list('id', flat=True)
        course = Course.objects.get(course_id=course_id, year=year, semester=semester)
        
        count =0 

        for student_id in student_ids:
            student = Student.objects.get(id=student_id)    
            count+=1
            # 이미 해당 수업에 등록된 학생이 있는지 확인
            if Course_registration.objects.filter(course=course, student=student).exists():
                print(f"{student} already registered for {course.name}")
                continue
            
            course_reg = Course_registration.objects.create(
                course=course,  
                course_year=course.year,
                course_semester=course.semester,
                student=student
            )
            print(f"OK! {student} has been registered for {course.name}")
            
        print(f"The num of students:{count}")
        course.student_count = count
        course.save()
        return JsonResponse({"status": "OK", "message": "course_registration record created successfully"})
    
    except Exception as e:
        return JsonResponse({"status": "Error", "message": str(e)}, status=500)

    
def course_project_update(request):
    try:

        courses = Course.objects.all()

        for course in courses:
            repos = Repository.objects.filter(name__icontains=course.course_repo_name)

            for repo in repos:
                # 이미 해당 course_id, year, semester, repo_id를 가지는 레코드가 있는지 확인
                if Course_project.objects.filter(course=course, course_year=course.year, course_semester=course.semester, repo_id=repo.id).exists():
                    print(f"{repo.owner_github_id}'s {repo.name} Course projects already exists!")
                    continue
                
                # 존재하지 않는다면 새로운 레코드 생성
                Course_project.objects.create(  
                    course=course,
                    course_year=course.year,
                    course_semester=course.semester,
                    repo_id=repo.id,
                    repo_name=repo.name
                )
                print(f"{repo.owner_github_id}'s {repo.name} has been created!")

        
        return JsonResponse({"status": "OK", "message": "course_project record created successfully"})
    except Exception as e:
        return JsonResponse({"status": "Error", "message": str(e)}, status=500)

def course_year_search(request):
    try:
        data = []
        year = request.GET.get('year')
        students = Student.objects.all()
        
        for student in students:
            # 특정 학생의 모든 레포지토리 가져옴
            student_repos = Repository.objects.filter(owner_github_id=student.github_id)
            
            # 특정 학생의 과목 관련된 레포지토리들을 가져옴
            courses_repos = Course_project.objects.filter(repo__in=student_repos)
            
            # 특정 학생이 듣는 모든 course_id 테이블 및 딕셔너리 생성
            course_ids = []
            for course_project in courses_repos:
                course_id = course_project.course.course_id
                if course_id not in course_ids:
                    course_ids.append(course_id)

            # 딕셔너리 초기화
            course_dict = {course_id: {'commit': 0, 'pr': 0, 'issue': 0, 'repo': 0} for course_id in course_ids}
            etc_dict = {year: {'commit': 0, 'pr': 0, 'issue': 0, 'repo': 0}}

            courses_project_repos = [c.repo for c in courses_repos]
            
            print(courses_project_repos)

            # 각 repo에 대한 정보 추가
            for repo in student_repos:
                repo_created_year = str(datetime.strptime(repo.created_at, '%Y-%m-%dT%H:%M:%SZ').year)  # 연도 추출
                if repo_created_year == year:
                    
                    if repo in courses_project_repos:  # 특정 학생의 repo가 과목과 관련이 있는 경우
                        specific_course = Course_project.objects.get(repo=repo)
                        course_id = specific_course.course.course_id
                    
                        # Commit 수 합산
                        course_dict[course_id]['commit'] += repo.commit_count
                        
                        # PR 수 합산
                        course_dict[course_id]['pr'] += Repo_pr.objects.filter(repo_id=repo.id).count()
                        
                        # Issue 수 합산
                        course_dict[course_id]['issue'] += Repo_issue.objects.filter(repo_id=repo.id).count()
                        
                        # Repo 수 합산
                        course_dict[course_id]['repo'] += 1
                    else:  # 특정 학생의 repo가 과목과 관련 없는 경우
                        etc_dict[repo_created_year]['commit'] += repo.commit_count
                        etc_dict[repo_created_year]['pr'] += Repo_pr.objects.filter(repo_id=repo.id).count()
                        etc_dict[repo_created_year]['issue'] += Repo_issue.objects.filter(repo_id=repo.id).count()
                        etc_dict[repo_created_year]['repo'] += 1

                else: 
                    continue 
            # 각 과목에 대한 정보 추가
            
        for course_id, course_count in course_dict.items():
                course = Course.objects.get(course_id=course_id)
                if str(course.year) == year :

                    course_info = {
                        "year": course.year,
                        "semester": course.semester,
                        "course_name": course.name,
                        "commit": course_count['commit'],
                        "pr": course_count['pr'],
                        "issue": course_count['issue'],
                        "num_repos": course_count['repo']
                    }

                    data.append(course_info)
        
        for the_year, the_year_count in etc_dict.items():
            etc_info = {
                "year": the_year,
                "semester": "",
                "course_name": the_year + " 기타",
                "commit": the_year_count['commit'],
                "pr": the_year_count['pr'],
                "issue": the_year_count['issue'],
                "num_repos": the_year_count['repo']
            }

            data.append(etc_info)               
                        
        return JsonResponse(data, safe=False)

    except Exception as e:
        return JsonResponse({"status": "Error", "message": str(e)}, status=500)
    

def course_department_count(request):
    try:
        data = []
        # 해당 년도의 모든 과목 코드들 가져옴
        
        # 간단하게 생각스.. 
        distinct_course_ids = Course.objects.all()
        print(distinct_course_ids)
        # Initialize the dictionary to store department counts

        for specific_course_id in distinct_course_ids:
            
            department_count = {}
            
            specific_course_reg = Course_registration.objects.filter(course = specific_course_id)

            for specific_course_per_reg in specific_course_reg : 
                student = Student.objects.get(id = specific_course_per_reg.student.id)
                
                # Get the department of the student
                department_name = student.department
                
                # Check if the department already exists in the dictionary
                if department_name in department_count:
                    # If it exists, increment the count
                    department_count[department_name] += 1
                else:
                    # If it doesn't exist, add it with a count of 1
                    department_count[department_name] = 1
            
            # You can now return this dictionary as part of the response, or process it further
            data.append({'year': specific_course_id.year,'semester': specific_course_id.semester ,'course_id': specific_course_id.course_id,'course_name': specific_course_id.name, 'department_count': department_count , "total_count": specific_course_id.student_count})
        
        return JsonResponse(data, safe=False)
    
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})

                


def course_validation(request):
    try:
        courses = Course.objects.all()

        for specific_course in courses:
            # 과목 학생 수 count validation 
            specific_course_count = Course_registration.objects.filter(course=specific_course).count()
            specific_course.student_count = specific_course_count
            print(f'{specific_course.name} {specific_course.course_id} student count has been changed to {specific_course.student_count}')
            specific_course.save()

        return JsonResponse({'status': 'success', 'message': 'Course student count updated'})

    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})


def course_reg_validation(request):
    try:
        access_token = get_kuopenapi_access_token()
        courses = Course.objects.all()
        data = []
        for specific_course in courses:
            course_id, course_class = specific_course.course_id.split('-')
            
            #API 요청
            api_url = "https://kuapi.korea.ac.kr/svc/course/course-registration/by-class"  # 실제 API 엔드포인트로 변경
            headers = {
                'AUTH_KEY': access_token
            }
            
            # 요청 파라미터 설정
            params = {
                'client_id': settings.KOREAUNIV_OPENAPI_CLIENT_ID,
                "year":str(specific_course.year),
                'term':str(specific_course.semester)+'R',
                'course_code': course_id,
                'course_class': course_class
            }
            
            # API 호출
            response = requests.get(api_url, headers=headers, params=params)
            # JSON 응답을 파싱
            
            response_data = response.json()
            
            result_items = response_data.get("result", [])
            
            new_ids = []

            for result_item in result_items:

                student_reg_id = result_item.get("student_id")
                
                try:
                    # Student 객체를 조회
                    reg_student = Student.objects.get(id=student_reg_id)
                    
                    try:
                        # Course_registration 객체가 존재하는지 확인
                        Course_registration.objects.get(student=reg_student, course = specific_course)
                        # 존재하면 pass
                        pass
                    
                    except Course_registration.DoesNotExist:
                        # Course_registration 객체가 존재하지 않으면 student_reg_id를 추가
                        Course_registration.objects.create(course = specific_course, course_year = specific_course.year , course_semester = specific_course.semester , student = reg_student)
                
                except Student.DoesNotExist:
                    # Student 객체가 존재하지 않으면 student_reg_id를 추가
                    new_ids.append(student_reg_id)

            
            #새로운 (미등록)학생들에 대해서 학생DB에 추가
            new_student_data,empty_ids = query_student_openapi(new_ids,access_token)
            print([specific_course.name,new_student_data,"empty",empty_ids])
            new_count = 0 

            for item in new_student_data :
                try :
                    id = item.get("STD_ID")
                    enrollment = item.get("REC_STS_NM")
                    name = item.get("KOR_NM")
                    college = item.get("COL_NM")
                    department = item.get("DEPT_NM")
                    double_major = item.get("SMAJOR_NM")
                    email_addr = item.get("email_addr")

                    new_reg_student, created = Student.objects.get_or_create(
                        id=id,                   
                        enrollment =enrollment,
                        name = name,
                        college = college,
                        department = department,
                        double_major = double_major,
                        primary_email = email_addr
                        
                    )
                    if not created:
                        print(f'Student with ID {id} already exists!')                    
                    
                    # Course_registration 객체 생성 또는 가져오기
                    new_course_reg, created = Course_registration.objects.get_or_create(
                        course=specific_course,
                        course_year=specific_course.year,
                        course_semester=specific_course.semester,
                        student=new_reg_student )
                    
                    if created:
                                print(f'Student {id}, {name} registered for {specific_course.name}')
                                new_count += 1
                    else:
                        print(f'Student {id}, {name} already registered for {specific_course.name}')

                except Exception as e:
                    print(f'Error processing student with ID {id}: {e}')
                    continue

            data.append(f'{specific_course.name}: new student created: {new_count} , empty student: {empty_ids}')

        return JsonResponse(data , safe=False) 

    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})



def course_reg_look(request):
    try:
        access_token = get_kuopenapi_access_token()
        courses = Course.objects.all()
        data = []
        for specific_course in courses:
            course_id, course_class = specific_course.course_id.split('-')
            
            #API 요청
            api_url = "https://kuapi.korea.ac.kr/svc/course/course-registration/by-class"  # 실제 API 엔드포인트로 변경
            headers = {
                'AUTH_KEY': access_token
            }
            
            # 요청 파라미터 설정
            params = {
                'client_id': settings.KOREAUNIV_OPENAPI_CLIENT_ID,
                "year":str(specific_course.year),
                'term':str(specific_course.semester)+'R',
                'course_code': course_id,
                'course_class': course_class
            }
            
            # API 호출
            response = requests.get(api_url, headers=headers, params=params)
            # JSON 응답을 파싱
            
            response_data = response.json()

            data.append(response_data)

        return JsonResponse(data , safe=False) 

    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})
    

def course_read_min_max_avg(request):
    try:
        courses = (
            Course.objects
            .select_related()
            .prefetch_related(
                Prefetch(
                    'course_project_set',
                    queryset=Course_project.objects.select_related('repo').filter(repo__forked=False)
                ),
                Prefetch(
                    'course_registration_set',
                    queryset=Course_registration.objects.select_related('student')
                )
            )
        )

        merged_stats = []
        total_stats_by_semester = defaultdict(lambda: {
            "total_student_count_sum": 0,
            "total_repository_count_sum": 0,
            "total_commits_sum": 0,
            "total_stars_sum": 0,
            "total_contributors_sum": 0,
            "course_count_in_semester": 0
        })

        for course in courses:
            course_key = (course.course_id, course.year, course.semester)

            course_projects = [
                cp for cp in course.course_project_set.all()
                if cp.repo and not cp.repo.forked
            ]
            repo_count = len(course_projects)
            total_commits = sum((cp.repo.commit_count or 0) for cp in course_projects)
            total_issues = sum(
                (cp.repo.open_issue_count or 0) + (cp.repo.closed_issue_count or 0)
                for cp in course_projects
            )
            total_prs = sum(
                (cp.repo.open_pr_count or 0) + (cp.repo.closed_pr_count or 0)
                for cp in course_projects
            )
            total_stars = sum((cp.repo.star_count or 0) for cp in course_projects)

            student_regs = list(course.course_registration_set.all())
            student_count = len(student_regs)

            contributor_count = 0
            repos_by_owner = defaultdict(list)
            for cp in course_projects:
                if cp.repo.contributors:
                    contributors = [c.strip() for c in cp.repo.contributors.split(',') if c.strip()]
                    contributor_count += len(contributors)
                repos_by_owner[cp.repo.owner_github_id].append(cp)

            student_stats = {'commits': [], 'prs': [], 'issues': [], 'num_repos': [], 'stars': []}
            for reg in student_regs:
                repos_for_student = repos_by_owner.get(reg.student.github_id, [])
                if not repos_for_student:
                    continue
                student_stats['commits'].append(sum(cp.repo.contributed_commit_count or 0 for cp in repos_for_student))
                student_stats['prs'].append(sum(
                    (cp.repo.contributed_open_pr_count or 0) + (cp.repo.contributed_closed_pr_count or 0)
                    for cp in repos_for_student
                ))
                student_stats['issues'].append(sum(
                    (cp.repo.contributed_open_issue_count or 0) + (cp.repo.contributed_closed_issue_count or 0)
                    for cp in repos_for_student
                ))
                student_stats['num_repos'].append(len(repos_for_student))
                student_stats['stars'].append(sum(cp.repo.star_count or 0 for cp in repos_for_student))

            def calc(values):
                if not values:
                    return {'q1': 0, 'q2': 0, 'q3': 0, 'std': 0, 'min': 0, 'max': 0}
                arr = np.array(values)
                return {
                    'q1': float(np.percentile(arr, 25)),
                    'q2': float(np.percentile(arr, 50)),
                    'q3': float(np.percentile(arr, 75)),
                    'std': float(np.std(arr)),
                    'min': float(np.min(arr)),
                    'max': float(np.max(arr))
                }

            commit_stats = calc(student_stats['commits'])
            pr_stats = calc(student_stats['prs'])
            issue_stats = calc(student_stats['issues'])
            num_repo_stats = calc(student_stats['num_repos'])
            star_stats = calc(student_stats['stars'])

            avg_commits = round(total_commits / student_count, 2) if student_count else 0

            merged_stats.append({
                "course_id": course.course_id,
                "year": course.year,
                "semester": course.semester,
                "course_name": course.name,
                "prof": course.prof,
                "ta": course.ta,
                "student_count": student_count,
                "total_commits": total_commits,
                "total_issues": total_issues,
                "total_prs": total_prs,
                "total_stars": total_stars,
                "avg_commits": avg_commits,
                "repository_count": repo_count,
                "contributor_count": contributor_count,
                "commit_min": commit_stats['min'],
                "commit_max": commit_stats['max'],
                "pr_min": pr_stats['min'],
                "pr_max": pr_stats['max'],
                "issue_min": issue_stats['min'],
                "issue_max": issue_stats['max'],
                "num_repos_min": num_repo_stats['min'],
                "num_repos_max": num_repo_stats['max'],
                "star_count_min": star_stats['min'],
                "star_count_max": star_stats['max'],
                "commit_q1": commit_stats['q1'],
                "commit_q2": commit_stats['q2'],
                "commit_q3": commit_stats['q3'],
                "commit_std": commit_stats['std'],
                "pr_q1": pr_stats['q1'],
                "pr_q2": pr_stats['q2'],
                "pr_q3": pr_stats['q3'],
                "pr_std": pr_stats['std'],
                "issue_q1": issue_stats['q1'],
                "issue_q2": issue_stats['q2'],
                "issue_q3": issue_stats['q3'],
                "issue_std": issue_stats['std'],
                "num_repos_q1": num_repo_stats['q1'],
                "num_repos_q2": num_repo_stats['q2'],
                "num_repos_q3": num_repo_stats['q3'],
                "num_repos_std": num_repo_stats['std'],
                "star_q1": star_stats['q1'],
                "star_q2": star_stats['q2'],
                "star_q3": star_stats['q3'],
                "star_std": star_stats['std'],
            })

            semester_key = (course.year, course.semester)
            total_stats_by_semester[semester_key]["total_student_count_sum"] += student_count
            total_stats_by_semester[semester_key]["total_repository_count_sum"] += repo_count
            total_stats_by_semester[semester_key]["total_commits_sum"] += total_commits
            total_stats_by_semester[semester_key]["total_stars_sum"] += total_stars
            total_stats_by_semester[semester_key]["total_contributors_sum"] += contributor_count
            total_stats_by_semester[semester_key]["course_count_in_semester"] += 1

        final_total_stats = []
        for (year, semester), stats in total_stats_by_semester.items():
            students = stats["total_student_count_sum"] or 1
            final_total_stats.append({
                "year": year,
                "semester": semester,
                "avg_repository_count": round(stats["total_repository_count_sum"] / students, 2),
                "avg_commits": round(stats["total_commits_sum"] / students, 2),
                "avg_stars": round(stats["total_stars_sum"] / students, 2),
                "avg_contributors": round(stats["total_contributors_sum"] / students, 2),
                "total_courses_in_semester": stats["course_count_in_semester"],
                "total_student_counts_in_semester": stats["total_student_count_sum"],
                "total_repository_counts_in_semester": stats["total_repository_count_sum"],
                "total_commits_in_semester": stats["total_commits_sum"],
                "total_stars_in_semester": stats["total_stars_sum"],
                "total_contributors_in_semester": stats["total_contributors_sum"],
            })

        return JsonResponse({
            "group_stats": merged_stats,
            "total_stats_by_semester": final_total_stats,
        }, safe=False)
    except Exception as e:
        import traceback
        print(f"Error in course_read_min_max_avg: {traceback.format_exc()}")
        return JsonResponse({"error": str(e)}, status=500)
    
def course_read_db_total_excel(request):
    try:
        data = []
        courses = Course.objects.all()

        for course in courses:

            total_commits = 0
            total_prs = 0
            total_issues = 0
            total_stars = 0

            # Gather course related repos' id
            course_related_repos_id = Course_project.objects.filter(course_id=course.id).values_list('repo_id', flat=True)
            print(f'{course.name}:[{course.course_id}] - repo_count: {len(course_related_repos_id)}')

            total_course_repos = Repository.objects.filter(id__in=course_related_repos_id)

            for course_repo in total_course_repos:
                if course_repo.forked is True:
                    continue
                else:
                    total_commits += course_repo.commit_count or 0
                    total_issues += ((course_repo.open_issue_count or 0) + (course_repo.closed_issue_count or 0))
                    total_prs += ((course_repo.open_pr_count or 0) + (course_repo.closed_pr_count or 0))
                    total_stars += course_repo.star_count or 0

            # Calculate the average commits
            student_count = Course_registration.objects.filter(
                course=course,
                course_year=course.year,
                course_semester=course.semester
            ).count()

            avg_commits = round(total_commits / student_count, 2) if student_count > 0 else 0

            # Calculate the number of repositories
            repository_count = Course_project.objects.filter(
                course=course,
                course_year=course.year,
                course_semester=course.semester
            ).count()

            # Calculate the number of contributors
            course_repo_ids = Course_project.objects.filter(
                course=course,
                course_year=course.year,
                course_semester=course.semester
            )

            contributor_count = 0
            for course_repo_id in course_repo_ids:
                if course_repo_id.repo.forked == False:
                    contributors_list_str = course_repo_id.repo.contributors
                    cleaned_contributors_list = ''.join(contributors_list_str.split())  # Remove all whitespace
                    if cleaned_contributors_list == '':
                        continue

                    contributors_number = cleaned_contributors_list.count(',') + 1  # Count contributors by commas
                    contributor_count += contributors_number

            # Gather all the data
            data.append({
                "course_id": course.course_id,
                "year": course.year,
                "semester": course.semester,
                "name": course.name,
                "prof": course.prof,
                "ta": course.ta,
                "student_count": course.student_count,
                "total_commits": total_commits,
                "total_issues": total_issues,
                "total_prs": total_prs,
                "total_stars": total_stars,
                "avg_commits": avg_commits,
                "repository_count": repository_count,
                "contributor_count": contributor_count
            })

        # Create Excel file
        output = BytesIO()
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Course Data"

        # Write headers
        headers = [
            "Course ID", "Year", "Semester", "Name", "Professor", "TA", "Student Count",
            "Total Commits", "Total Issues", "Total PRs", "Total Stars", "Average Commits",
            "Repository Count", "Contributor Count"
        ]
        sheet.append(headers)

        # Write data rows
        for row in data:
            sheet.append([
                row["course_id"], row["year"], row["semester"], row["name"], row["prof"], row["ta"],
                row["student_count"], row["total_commits"], row["total_issues"], row["total_prs"],
                row["total_stars"], row["avg_commits"], row["repository_count"], row["contributor_count"]
            ])

        # Adjust column widths
        for column_cells in sheet.columns:
            max_length = 0
            column = column_cells[0].column_letter  # Get the column name
            for cell in column_cells:
                try:  # Necessary to avoid issues with NoneType
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            adjusted_width = max_length + 2
            sheet.column_dimensions[column].width = adjusted_width

        # Save workbook to BytesIO
        workbook.save(output)
        output.seek(0)

        # Return as Excel file response
        response = HttpResponse(
            output,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response['Content-Disposition'] = 'attachment; filename="course_data.xlsx"'
        return response

    except Exception as e:
        return JsonResponse({"status": "Error", "message": str(e)}, status=500)

def course_student_db_excel(request):
    try:
        # 데이터를 가져오는 함수 호출
        response = student_read_course_info(request)

        # JsonResponse에서 데이터 추출
        if hasattr(response, 'content'):
            student_course_list = json.loads(response.content.decode('utf-8'))
        else:
            raise ValueError("Invalid response format from student_read_course_info")

        # 반환 데이터 형식 확인
        if not isinstance(student_course_list, list):
            raise ValueError("Invalid data format: Expected a list of dictionaries.")

    except Exception as e:
        return HttpResponse(f"Error processing student data: {str(e)}", status=500)

    # course_id, year, semester로 데이터 그룹화
    grouped_data = {}
    for student in student_course_list:
        try:
            course_id = student.get("course_id", "").strip() or "기타"
            year = student.get("year") or 0  # 기본값을 0으로 설정
            semester = student.get("semester") or 0  # 기본값을 0으로 설정

            # Ensure year and semester are integers
            year = int(year) if year else 0
            semester = int(semester) if semester else 0

            group_key = f"{course_id}_{year}_{semester}"
            if group_key not in grouped_data:
                grouped_data[group_key] = []
            
            # Repository Name 가져오기
            try:
                student_id = student.get("id")
                specific_student = Student.objects.get(id=student_id)
                course_info= Course.objects.get(course_id=course_id, year=year, semester=semester)
                course_repositories = Course_project.objects.filter(course = course_info)
                for course_repo in course_repositories:
                    if course_repo.repo.owner_github_id == specific_student.github_id :
                        repository_name = course_repo.repo.name
                        break
                    else :
                        continue
            except ObjectDoesNotExist:
                repository_name = "Unknown"

            student["repository_name"] = repository_name
            grouped_data[group_key].append(student)
        except AttributeError as e:
            return HttpResponse(f"Error processing student record: {str(e)}", status=500)

    # 엑셀 파일 생성
    output = io.BytesIO()
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)  # 기본 시트 제거

    for group_key, students in grouped_data.items():
        # 시트 이름 생성 (길이 제한 고려)
        sheet_name = group_key[:31]  # 시트 이름은 최대 31자
        sheet = workbook.create_sheet(title=sheet_name)

        # 헤더 작성
        headers = [
            "ID", "GitHub ID", "Name", "Department", "Enrollment",
            "Year", "Semester", "Course Name", "Commit", "PR", "Issue",
            "Num Repos", "Star Count", "Professor", "Total Contributors",  "Repository Names"
        ]
        sheet.append(headers)
        for col in sheet[1]:
            col.font = Font(bold=True)

        # 학생 데이터 삽입
        for student in students:
            row = [
                student.get("id"),
                student.get("github_id"),
                student.get("name"),
                student.get("department"),
                student.get("enrollment"),
                student.get("year"),
                student.get("semester"),
                student.get("course_name"),
                student.get("commit"),
                student.get("pr"),
                student.get("issue"),
                student.get("num_repos"),
                student.get("star_count"),
                student.get("prof"),
                student.get("total_contributors"),
                student.get("repository_name")
            ]
            sheet.append(row)

    # 엑셀 파일 저장
    workbook.save(output)
    output.seek(0)

    # HTTP 응답으로 엑셀 파일 반환
    response = HttpResponse(
        output,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename=\"course_students.xlsx\"'
    return response
# def course_student_db_excel(request):
#     try:
#         data = []
#         courses = Course.objects.all()

#         for course in courses:
            
#             course_id = course.course_id
#             year = course.year
#             semester = course.semester


#             # Gather course related repos' id
#             course_related_repos_id = Course_project.objects.filter(course_id=course.id).values_list('repo_id', flat=True)
#             print(f'{course.name}:[{course.course_id}] - repo_count: {len(course_related_repos_id)}')
            
            
#             total_course_repos = Repository.objects.filter(id__in=course_related_repos_id)

#             for course_repo in total_course_repos :

#                 if course_repo.forked is True :
#                     continue            
#                 else :
                    
#                     student = Student.objects.get(github_id = course_repo.owner_github_id)                
#                     student_id = student.id 
#                     student_github_id = student.github_id
#                     student_name = student.name 
#                     student_department = student.department
#                     student_enrollment = student.enrollment
                    
#                     commit_couunt = course_repo.commit_count or 0 
#                     issue_count =  ((course_repo.open_issue_count or 0) + (course_repo.closed_issue_count or 0))
#                     pr_count = ((course_repo.open_pr_count or 0) + (course_repo.closed_pr_count or 0))
#                     star_count = course_repo.star_count or 0 
            

#                     if course_repo.forked == False :
#                         contributors_list_str = course_repo.contributors
#                         cleaned_contributors_list = ''.join(contributors_list_str.split())  # 모든 공백 제거
#                         if cleaned_contributors_list == '':
#                             continue
#                         contributors_number = cleaned_contributors_list.count(',') + 1 # , 코마갯수로 contributor 세기

#     except Exception as e:
#         return JsonResponse({"status": "Error", "message": str(e)}, status=500)


def only_for_course_student_db_excel(request): #과목 관점에서 읽기 
    try:
        data = []
        students = Student.objects.all()
        for student in students:
            try:
                # Proceed for one specific student             
                total_commit = 0 
                total_pr = 0
                total_issue = 0
                total_repo = 0
                total_star = 0

                etc_total_commit = 0 
                etc_total_pr = 0
                etc_total_issue = 0
                etc_total_repo = 0
                etc_total_star = 0
                etc_total_contributor=0

                # Get all repositories for a specific student
                student_repos = Repository.objects.filter(owner_github_id=student.github_id)
                
                # 특정 학생의 수강 목록들 가져옴
                course_reg_list = Course_registration.objects.filter(student=student)

                # 특정 학생의 과목 관련된 레포지토리들을 가져옴.
                courses_repos = Course_project.objects.filter(repo__in=student_repos)
                
                # 특정 학생이 듣는 모든 course_id 테이블 및 딕셔너리 생성
                course_ids = []
                for course_reg in course_reg_list:
                    course_ids.append(course_reg.course.course_id)

                # 딕셔너리 초기화
                course_dict = {course_id: {'commit': 0, 'pr': 0, 'issue': 0, 'repo': 0, 'star': 0, 'contributors': 0,} for course_id in course_ids}

                courses_project_repos = []

                for c in courses_repos:
                    courses_project_repos.append(c.repo)

                # 각 repo에 대한 정보 추가
                for repo in student_repos:
                    sanitized_url = repo.url.replace(":", "")

                    if repo in courses_project_repos:  # 특정 학생의 현재 가리키는 repo가 과목과 관련이 있는 경우
                        specific_course = Course_project.objects.get(repo=repo)
                        course_id = specific_course.course.course_id

                        # Commit 수 합산
                        total_commit_count = Repository.objects.filter(
                            id = specific_course.repo.id,
                            owner_github_id=student.github_id
                        ).aggregate(total_commits=Sum('contributed_commit_count'))['total_commits'] or 0
                        
                        # 합산한 값을 course_dict에 추가
                        course_dict[course_id]['commit'] += total_commit_count
                        
                        # PR 수 합산
                        course_dict[course_id]['pr'] += ( repo.contributed_open_pr_count or 0) +  (repo.contributed_closed_pr_count or 0)
                        
                        # Issue 수 합산
                        course_dict[course_id]['issue'] += (repo.contributed_open_issue_count or 0) + (repo.contributed_closed_issue_count or 0)
                        
                        # Repo 수 합산
                        course_dict[course_id]['repo'] += 1 
                        
                        # Star 수 합산
                        course_dict[course_id]['star'] += Repository.objects.get(id=repo.id).star_count or 0
                        print(repo.url)

                        course_dict[course_id]['contributors'] += Repo_contributor.objects.filter(repo_url = sanitized_url).count() or 0 

                    else:  # 특정 학생의 repo가 과목과 관련 없는 경우
                        etc_total_commit += repo.contributed_commit_count or 0 
                        etc_total_pr +=  ( repo.contributed_open_pr_count or 0) +  (repo.contributed_closed_pr_count or 0)
                        etc_total_issue += (repo.contributed_open_issue_count or 0) + (repo.contributed_closed_issue_count or 0)
                        etc_total_repo += 1
                        etc_total_star += Repository.objects.get(id=repo.id).star_count or 0
                        etc_total_contributor += Repo_contributor.objects.filter(repo_url = sanitized_url).count() or 0

                # 각 과목에 대한 정보 추가
                for course_id, course_count in course_dict.items():
                    course = Course.objects.get(course_id=course_id)

                    course_info = {
                        "id": student.id,
                        "github_id": student.github_id,
                        "name": student.name,
                        "department": student.department,
                        "enrollment": student.enrollment,
                        "year": course.year,
                        "semester": course.semester,
                        "course_name": course.name,
                        "commit": course_count['commit'],
                        "pr": course_count['pr'],
                        "issue": course_count['issue'],
                        "num_repos": course_count['repo'],
                        "star_count": course_count['star'],
                        "prof": course.prof,
                        "course_id": course.course_id,
                        "total_contributors": course_count['contributors']
                    }
                    total_commit += course_count['commit']
                    total_pr += course_count['pr']
                    total_issue += course_count['issue']
                    total_repo += course_count['repo']
                    total_star += course_count['star']

                    data.append(course_info)

                # 기타 정보 추가
                etc_info = {
                    "id": student.id,
                    "github_id": student.github_id,
                    "name": student.name,
                    "department": student.department,
                    "enrollment": student.enrollment,
                    "year": "",
                    "semester": "",
                    "course_name": "기타",
                    "commit": etc_total_commit,
                    "pr": etc_total_pr,
                    "issue": etc_total_issue,
                    "num_repos": etc_total_repo,
                    "star_count": etc_total_star,
                    "prof": "",
                    "course_id": "",
                    "total_contributors": etc_total_contributor
                }

                total_commit += etc_total_commit
                total_pr += etc_total_pr
                total_issue += etc_total_issue
                total_repo += etc_total_repo
                total_star += etc_total_star

                data.append(etc_info)               

            except Exception as e:
                print(f'Error processing student {student.name}: {e}')
                continue       

        return JsonResponse(data, safe=False)

    except Exception as e:
        return JsonResponse({"status": "Error", "message": str(e)}, status=500)